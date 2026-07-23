from datetime import timedelta
from decimal import Decimal
from io import BytesIO

import pytest
from django.utils import timezone
from openpyxl import load_workbook

from apps.accounts.models import User
from apps.bookings.models import BookableSpace
from apps.events.models import Event
from apps.machines.models import Machine, MachineType, MachineUsageEntry
from apps.operations.report_registry import REPORT_REGISTRY
from tests.return_helpers import authenticated_client, make_member, make_space, make_user


pytestmark = pytest.mark.django_db
KEYS = ("machine-usage", "event-attendance", "booking-utilization", "maintenance-activity", "fablab-health")


def _seed(slug):
    space = make_space(slug)
    manager = make_member(f"{slug}-manager", space)
    machine_type = MachineType.objects.create(makerspace=space, slug=f"{slug}-type", name="Safe type")
    machine = Machine.objects.create(makerspace=space, machine_type=machine_type, name="Safe machine")
    MachineUsageEntry.objects.create(machine=machine, hours=Decimal("1.25"))
    Event.objects.create(makerspace=space, title="Safe event", starts_at=timezone.now(), ends_at=timezone.now() + timedelta(hours=1))
    BookableSpace.objects.create(makerspace=space, name="Safe space")
    return space, manager


@pytest.mark.parametrize("key", KEYS)
@pytest.mark.parametrize("fmt", ["csv", "xlsx"])
def test_each_new_key_exports_exact_headers_per_space_and_aggregate(key, fmt):
    space, manager = _seed(f"export-{key}-{fmt}")
    per = authenticated_client(manager).get(
        f"/api/v1/admin/makerspace/{space.id}/reports/{key}/export?format={fmt}&limit=1"
    )
    assert per.status_code == 200
    assert _header(per, fmt) == list(REPORT_REGISTRY[key].fields)

    disabled, _ = _seed(f"export-disabled-{key}-{fmt}")
    disabled.enabled_modules = [module for module in disabled.enabled_modules if module != "reports"]
    disabled.save(update_fields=["enabled_modules"])

    superadmin = make_user(
        f"export-super-{key}-{fmt}", role=User.Role.SUPERADMIN,
        access_status=User.AccessStatus.ACTIVE,
    )
    aggregate = authenticated_client(superadmin).get(
        f"/api/v1/admin/reports/{key}/export?format={fmt}"
    )
    assert aggregate.status_code == 200
    assert _header(aggregate, fmt) == ["makerspace_id", *REPORT_REGISTRY[key].fields]
    ids = _makerspace_ids(aggregate, fmt)
    assert space.id in ids
    assert disabled.id not in ids


def test_exports_keep_decimal_numeric_and_escape_formula_cells():
    space, manager = _seed("export-formula")
    event = Event.objects.filter(makerspace=space).first()
    event.title = "=1+1"
    event.save(update_fields=["title"])
    client = authenticated_client(manager)
    csv_response = client.get(f"/api/v1/admin/makerspace/{space.id}/reports/event-attendance/export")
    assert "'=1+1" in csv_response.content.decode()
    xlsx_response = client.get(f"/api/v1/admin/makerspace/{space.id}/reports/event-attendance/export?format=xlsx")
    values = _values(xlsx_response, "xlsx")
    assert "'=1+1" in values
    usage = client.get(f"/api/v1/admin/makerspace/{space.id}/reports/machine-usage/export?format=xlsx")
    sheet = load_workbook(BytesIO(usage.content)).active
    assert sheet.cell(2, 6).data_type == "n"
    assert Decimal(str(sheet.cell(2, 6).value)) == Decimal("1.25")


def test_json_request_invokes_canonical_builder_once(monkeypatch):
    space, manager = _seed("export-builder-once")
    from apps.operations import reports_machine_usage

    original = reports_machine_usage.build_machine_usage
    calls = 0

    def counted(*args, **kwargs):
        nonlocal calls
        calls += 1
        return original(*args, **kwargs)

    monkeypatch.setattr(reports_machine_usage, "build_machine_usage", counted)
    response = authenticated_client(manager).get(
        f"/api/v1/admin/makerspace/{space.id}/analytics/machine-usage"
    )
    assert response.status_code == 200
    assert calls == 1
    assert response.data["rows"][1][5] == response.data["typed_rows"][0]["usage_hours"]


def test_export_helper_compatibility_reexports_remain_live():
    from apps.operations import views, views_ledger, views_reports

    assert views._csv_response is views_reports._csv_response
    assert views._xlsx_response is views_reports._xlsx_response
    assert views_ledger._csv_response is views_reports._csv_response
    assert views_ledger._xlsx_response is views_reports._xlsx_response
    assert views._xlsx_cell is views_reports._xlsx_cell


def _header(response, fmt):
    if fmt == "csv":
        return response.content.decode().splitlines()[0].split(",")
    return [cell.value for cell in load_workbook(BytesIO(response.content)).active[1]]


def _values(response, fmt):
    if fmt == "csv":
        return response.content.decode()
    return repr([[cell.value for cell in row] for row in load_workbook(BytesIO(response.content)).active.iter_rows()])


def _makerspace_ids(response, fmt):
    if fmt == "csv":
        return {int(line.split(",", 1)[0]) for line in response.content.decode().splitlines()[1:]}
    rows = load_workbook(BytesIO(response.content)).active.iter_rows(min_row=2, values_only=True)
    return {int(row[0]) for row in rows}
